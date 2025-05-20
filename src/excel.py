from itertools import product
from pathlib import Path
import shutil
import os
import re
import math

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config_loader import Config, load_column_names


def apply_cell_style(cell, font, fill=None, alignment=None, border=None):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def estimate_column_width(text, font_name='Meiryo UI', font_size=11, min_width=8, max_width=100):
    """
    テキストの内容に基づいて適切な列幅を推定します。
    日本語文字は英数字より幅が広いため、文字数ベースで計算します。
    
    Args:
        text (str): セルに表示するテキスト
        font_name (str): フォント名
        font_size (int): フォントサイズ
        min_width (int): 最小列幅
        max_width (int): 最大列幅
        
    Returns:
        float: 推定列幅
    """
    if not text:
        return min_width
    
    # 改行で分割して各行の長さを計算
    lines = str(text).split('\n')
    
    # 各行の文字数に基づいて幅を計算
    widths = []
    for line in lines:
        # 全角文字（主に日本語）は2文字分、半角は1文字分として計算
        # 全角文字のパターン
        fullwidth_pattern = re.compile(r'[^\x00-\xff]')
        
        # 全角文字の数を数える
        fullwidth_count = len(fullwidth_pattern.findall(line))
        
        # 半角文字の数を数える
        halfwidth_count = len(line) - fullwidth_count
        
        # 文字数ベースの幅を計算（全角:半角 = 2:1の比率）
        char_width = fullwidth_count * 2 + halfwidth_count
        
        # フォントサイズに応じて調整（Excelの列幅単位に変換）
        # 標準のフォントサイズ11ptを基準に計算
        excel_width = char_width * (font_size / 11) * 1.5  # 余裕係数を1.2から1.5に増加
        
        widths.append(excel_width)
    
    # 最大幅を求める
    estimated_width = max(widths)
    
    # 最小・最大の範囲内に収める
    return max(min_width, min(estimated_width, max_width))


def estimate_row_height(row_data, font_name='Meiryo UI', font_size=11, min_height=15, line_height_factor=1.5):
    """
    行データの内容に基づいて適切な行の高さを推定します。
    
    Args:
        row_data (list): 行データ（各セルの値のリスト）
        font_name (str): フォント名
        font_size (int): フォントサイズ
        min_height (float): 最小行高
        line_height_factor (float): 行間の余裕係数
        
    Returns:
        float: 推定行高
    """
    max_lines = 1
    
    # 各セルの改行数を計算
    for value in row_data:
        if value is None:
            continue
            
        # テキストを文字列に変換して改行で分割
        lines = str(value).split('\n')
        
        # 各行の文字数によって必要な折り返し回数を考慮
        wrapped_line_count = 0
        for line in lines:
            # 全角文字（主に日本語）の場合、1行あたりの文字数は少なくなる
            # 全角文字のパターン
            fullwidth_pattern = re.compile(r'[^\x00-\xff]')
            
            # 全角文字の数を数える
            fullwidth_count = len(fullwidth_pattern.findall(line))
            
            # 半角文字の数を数える
            halfwidth_count = len(line) - fullwidth_count
            
            # 文字数ベースの幅を計算（全角:半角 = 2:1の比率）
            char_width = fullwidth_count * 2 + halfwidth_count
            
            # 1行あたり約40-50文字として折り返し回数を概算
            # これは列幅に依存するため、厳密な計算は難しい
            # 余裕を見て少し多めに見積もる
            needed_lines = max(1, char_width / 40)
            wrapped_line_count += needed_lines
        
        max_lines = max(max_lines, wrapped_line_count)
    
    # 行の高さを計算（各行に対して余裕を持たせる）
    # Excel行の高さの単位はポイント（pt）なので、フォントサイズに比例
    # 小数点以下は切り上げて、より余裕を持たせる
    estimated_height = math.ceil(max_lines) * font_size * line_height_factor
    
    return max(min_height, estimated_height)


class ExcelWriter:

    def __init__(self, df: pd.DataFrame, config_excel: Config):
        self.df = df.copy()
        self.config = config_excel

        self.columns = load_column_names(self.config)
        self.col_names = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:len(self.columns)]
        if len(self.col_names) < len(self.columns):
            raise IndexError("列は26列以下にしてください")

    def __write_test_specification_sheet(self,
                                         workbook,
                                         merge_cells: bool = False,
                                         template_used: bool = False,
                                         auto_adjust_width: bool = True,
                                         auto_adjust_height: bool = True,
                                         preserve_additional_columns: bool = False,
                                         ):
        """テスト仕様書をエクセルシートに書き込みます。

        Args:
            workbook:               openpyxlのワークブックオブジェクト
            merge_cells (bool):     セルをマージするかどうか
            template_used (bool):   テンプレートを使用しているかどうか
            auto_adjust_width (bool): 列幅を内容に合わせて自動調整するかどうか
            auto_adjust_height (bool): 行高を内容に合わせて自動調整するかどうか
            preserve_additional_columns (bool): J列以降の内容を保持するかどうか
        """
        df_excel = self.df.copy()
        df_excel.columns = self.columns

        # マージできるようにマルチインデックス化
        multi_idx_cols = []
        if merge_cells:
            multi_idx_cols = [v["name"] for k, v in self.config.columns.model_dump().items() if v.get("multi_idx")]
            if multi_idx_cols:
                # マージセルを使うが、NOをA列に確実に表示するため別の方法でマージする
                pass
            else:
                merge_cells = False  # マージ対象の列がなければマージを無効化

        # シート取得
        sheet_name = self.config.excel_settings.sheet_name.test
        worksheet = workbook[sheet_name] if template_used else workbook.create_sheet(sheet_name)

        # テンプレートを使用している場合、既存のシートにデータを追加する
        if template_used:
            # 既存データの最終行を取得 (ヘッダー行を考慮)
            last_row = 1
            for row in worksheet.iter_rows():
                if row[0].value is not None:
                    last_row += 1
                else:
                    break
            
            # J列以降のデータを保存する（オプションが有効な場合）
            additional_columns_data = {}
            if preserve_additional_columns:
                # J列から右側のデータを一時保存 (通常、columnのインデックスは0から始まるため、J列はインデックス9)
                start_col_idx = 9  # J列
                for row_idx in range(2, last_row):  # 2行目から最終行まで（ヘッダー行はスキップ）
                    row_data = {}
                    # 行に一意の識別子を付ける（A列の値）
                    row_id = worksheet["A" + str(row_idx)].value
                    if row_id is not None:
                        for col_idx in range(start_col_idx, worksheet.max_column + 1):
                            col_letter = get_column_letter(col_idx + 1)
                            cell_value = worksheet[col_letter + str(row_idx)].value
                            row_data[col_idx] = cell_value
                        additional_columns_data[row_id] = row_data
            
            # データフレームの内容をシートに書き込む (ヘッダー行はスキップ)
            for i, row in enumerate(df_excel.itertuples(index=False)):
                # 行の高さを自動調整
                if auto_adjust_height:
                    row_height = estimate_row_height(row, self.config.excel_settings.font_name)
                    worksheet.row_dimensions[last_row + i].height = row_height
                
                for j, value in enumerate(row):
                    col_letter = get_column_letter(j+1)  # 1-indexed
                    cell = worksheet[f"{col_letter}{last_row + i}"]
                    cell.value = value
                    
                    # スタイル適用
                    apply_cell_style(
                        cell,
                        font=Font(name=self.config.excel_settings.font_name),
                        alignment=Alignment(
                            horizontal=self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]["horizontal"], 
                            vertical=self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]["vertical"], 
                            wrap_text=True
                        ),
                        border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                    )
                
                # J列以降のデータを復元（オプションが有効な場合）
                if preserve_additional_columns:
                    row_id = row[0]  # A列（NOカラム）の値
                    if row_id in additional_columns_data:
                        for col_idx, cell_value in additional_columns_data[row_id].items():
                            col_letter = get_column_letter(col_idx + 1)
                            cell = worksheet[f"{col_letter}{last_row + i}"]
                            if cell_value is not None:  # None以外の値のみ設定
                                cell.value = cell_value
            
            # G列からM列まで（試験実施者から再試験結果備考まで）の枠線を追加
            for row_idx in range(last_row, last_row + len(df_excel)):
                for col_idx in range(7, 14):  # G列(7)からM列(13)まで
                    col_letter = get_column_letter(col_idx)
                    cell = worksheet[f"{col_letter}{row_idx}"]
                    # スタイルのみ適用（枠線と文字の折り返し）
                    apply_cell_style(
                        cell,
                        font=Font(name=self.config.excel_settings.font_name),
                        alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
                        border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                    )
                    
                # J列以降の追加列の枠線も適用（読み込んだデータに基づく）
                if preserve_additional_columns and worksheet.max_column > 13:
                    for col_idx in range(14, worksheet.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        # スタイルのみ適用（枠線と文字の折り返し）
                        apply_cell_style(
                            cell,
                            font=Font(name=self.config.excel_settings.font_name),
                            alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
                            border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                        top=Side(style="thin"), bottom=Side(style="thin"))
                        )
        else:
            # 新規シートにデータを書き込む
            # ヘッダーを書き込む
            for j, col_name in enumerate(self.columns):
                col_letter = get_column_letter(j+1)  # 1-indexed
                cell = worksheet[f"{col_letter}1"]
                cell.value = col_name
                
                # ヘッダースタイル適用
                apply_cell_style(
                    cell,
                    font=Font(name=self.config.excel_settings.font_name, bold=True, color="ffffff"),
                    fill=PatternFill(patternType="solid", fgColor="4f81bd"),
                    alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
                )
                
                # デフォルトの列幅を設定
                worksheet.column_dimensions[col_letter].width = self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]["length"]
            
            # テンプレート列のヘッダー（G列からM列）も同様に設定
            additional_headers = ['試験\n実施者', '試験日', '試験\nステータス', '試験結果備考', '再試験\n実施者', '再試験\nステータス', '再試験結果備考']
            for j, header in enumerate(additional_headers, 7):  # G列(7)から始める
                col_letter = get_column_letter(j)
                cell = worksheet[f"{col_letter}1"]
                cell.value = header
                
                # ヘッダースタイル適用
                apply_cell_style(
                    cell,
                    font=Font(name=self.config.excel_settings.font_name, bold=True, color="ffffff"),
                    fill=PatternFill(patternType="solid", fgColor="4f81bd"),
                    alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
                )
                
                # デフォルトの列幅を設定（標準的な幅を適用）
                worksheet.column_dimensions[col_letter].width = 15
            
            # データを書き込む
            for i, row in enumerate(df_excel.itertuples(index=False)):
                # 行の高さを自動調整
                if auto_adjust_height:
                    row_height = estimate_row_height(row, self.config.excel_settings.font_name)
                    worksheet.row_dimensions[i + 2].height = row_height
                
                for j, value in enumerate(row):
                    col_letter = get_column_letter(j+1)  # 1-indexed
                    cell = worksheet[f"{col_letter}{i + 2}"]
                    cell.value = value
                    
                    # スタイル適用
                    apply_cell_style(
                        cell,
                        font=Font(name=self.config.excel_settings.font_name),
                        alignment=Alignment(
                            horizontal=self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]["horizontal"], 
                            vertical=self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]["vertical"], 
                            wrap_text=True
                        ),
                        border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                    )
            
            # G列からM列まで（試験実施者から再試験結果備考まで）の枠線を追加
            for row_idx in range(2, 2 + len(df_excel)):
                for col_idx in range(7, 14):  # G列(7)からM列(13)まで
                    col_letter = get_column_letter(col_idx)
                    cell = worksheet[f"{col_letter}{row_idx}"]
                    # スタイルのみ適用（枠線と文字の折り返し）
                    apply_cell_style(
                        cell,
                        font=Font(name=self.config.excel_settings.font_name),
                        alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
                        border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                    )
        
        # 列幅の自動調整（オプションが有効な場合）
        if auto_adjust_width:
            # 各列ごとに最適な幅を計算
            for j, col_name in enumerate(self.columns):
                col_letter = get_column_letter(j+1)  # 1-indexed
                col_config = self.config.columns.model_dump()[list(self.config.columns.model_dump().keys())[j]]
                
                # この列の最大幅を計算（ヘッダーも含む）
                max_width = estimate_column_width(col_name, self.config.excel_settings.font_name)
                
                # データ行をチェック
                start_row = 2
                end_row = len(df_excel) + (1 if not template_used else last_row - 1)
                
                for row in range(start_row, end_row + 1):
                    actual_row = row if not template_used else row + last_row - 2
                    cell_value = worksheet[f"{col_letter}{actual_row}"].value
                    if cell_value:
                        estimated_width = estimate_column_width(cell_value, self.config.excel_settings.font_name)
                        max_width = max(max_width, estimated_width)
                
                # 計算した最大幅を設定（configで指定された幅よりも大きい場合のみ）
                default_width = col_config["length"]
                worksheet.column_dimensions[col_letter].width = max(default_width, max_width)
            
            # G列からM列の幅も自動調整
            for col_idx in range(7, 14):  # G列(7)からM列(13)まで
                col_letter = get_column_letter(col_idx)
                # ヘッダーの幅を計算
                header_cell = worksheet[f"{col_letter}1"]
                max_width = estimate_column_width(header_cell.value, self.config.excel_settings.font_name)
                
                # デフォルト幅を設定（少なくとも12以上）
                worksheet.column_dimensions[col_letter].width = max(12, max_width)
                
        # マージセルの処理（テンプレート使用の有無にかかわらず適用）
        if merge_cells and multi_idx_cols:
            # マージ対象の列のインデックスを取得
            merge_col_indices = [self.columns.index(col) for col in multi_idx_cols]
            
            # 各マージ対象列に対して処理
            for col_idx in merge_col_indices:
                col_letter = get_column_letter(col_idx+1)  # 1-indexed
                
                # マージするセルの範囲を決定
                current_value = None
                start_row = 2 if not template_used else last_row  # データの開始行
                
                # 処理対象の行数
                row_count = len(df_excel)
                
                for i in range(row_count):
                    # 実際のExcel行番号
                    row = i + (2 if not template_used else last_row)
                    cell_value = worksheet[f"{col_letter}{row}"].value
                    
                    # 値が変わったか、最後の行の場合
                    if cell_value != current_value or i == row_count - 1:
                        # 前のグループがあればマージ
                        if current_value is not None and row - start_row > 1:
                            # 最後の行の場合は現在の行も含める
                            end_row = row if cell_value != current_value else row + 1
                            worksheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row-1}")
                        
                        # 新しいグループの開始
                        current_value = cell_value
                        start_row = row

    def __call__(self, output_path: Path, merge_cells: bool = True, template_path: Path = None, 
                auto_adjust_width: bool = True, auto_adjust_height: bool = True, preserve_additional_columns: bool = False):
        """
        convert_md_to_df()により生成されたデータフレームをエクセルファイルに変換します。

        Args:
            output_path (Path):        出力先のパス
            merge_cells (bool):        セルをマージするかどうか
            template_path (Path):      テンプレートとして使用するExcelファイルのパス。指定された場合はテンプレートをコピーして使用
            auto_adjust_width (bool):  列幅を内容に合わせて自動調整するかどうか
            auto_adjust_height (bool): 行高を内容に合わせて自動調整するかどうか
            preserve_additional_columns (bool): J列以降の内容を保持するかどうか（テンプレート使用時のみ有効）
        """
        try:
            # テンプレートが指定されている場合
            if template_path and template_path.exists():
                # テンプレートファイルと出力先が同じ場合は直接編集
                if template_path == output_path:
                    workbook = load_workbook(template_path)
                else:
                    # テンプレートファイルをコピー
                    shutil.copy2(template_path, output_path)
                    # コピーしたファイルを開く
                    workbook = load_workbook(output_path)
                
                # シートが存在することを確認
                sheet_name = self.config.excel_settings.sheet_name.test
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"テンプレートに'{sheet_name}'シートが見つかりません。")
                
                # テンプレートのシートにデータを書き込む
                self.__write_test_specification_sheet(workbook, merge_cells, template_used=True, 
                                                     auto_adjust_width=auto_adjust_width,
                                                     auto_adjust_height=auto_adjust_height,
                                                     preserve_additional_columns=preserve_additional_columns)
                
                # 変更を保存
                workbook.save(output_path)
            else:
                # 新規ファイルを作成
                workbook = load_workbook()
                
                # テスト仕様書シートを作成して書き込む
                self.__write_test_specification_sheet(workbook, merge_cells, template_used=False, 
                                                     auto_adjust_width=auto_adjust_width,
                                                     auto_adjust_height=auto_adjust_height,
                                                     preserve_additional_columns=False)  # 新規ファイルの場合はデータ保持は無意味
                
                # 不要なSheetを削除して保存
                if "Sheet" in workbook.sheetnames:
                    del workbook["Sheet"]
                
                workbook.save(output_path)
                    
        except PermissionError:
            raise PermissionError(f"出力先のファイルを開いている可能性があります。エクセルファイルを閉じてください。")
        except Exception as e:
            raise ValueError(f"エクセルファイル出力中に不明なエラーが発生しました：\n{e}")

        return output_path
